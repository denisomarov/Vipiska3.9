# объявляем библиотеки

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS
from flask import send_file

from datetime import date

# Процедура обработки выписки
def v_treatment(HalykFileName, SberFileName):

    # считываем exel файл по Сбербанку

    data_sber = pd.read_excel(SberFileName, sheet_name=0, header=6)

    # проверяем, есть ли записи об операциях в файле

    if len(data_sber) > 0:

        # удаляем последние 2 строки
        data_sber = data_sber.loc[:(len(data_sber) - 3)]

        # конверируем дату в формат datetime
        data_sber['Дата валютирования'] = pd.to_datetime(data_sber['Дата валютирования'], format='%d.%m.%Y')

    # считываем exel файл по Halyk банку
    data_halyk = pd.read_excel(HalykFileName, sheet_name=0, header=10)

    # проверяем, есть ли записи об операциях в файле

    if len(data_halyk) > 0:
        # удаляем последние 2 строки
        data_halyk = data_halyk.loc[:(len(data_halyk) - 3)]

        # конверируем дату в формат datetime
        data_halyk['Дата валютирования'] = pd.to_datetime(data_halyk['Дата валютирования'], format='%d.%m.%Y')

        # дополняем в таблицу Халык Банк столбец Дата
        data_halyk.insert(0, 'Дата', '')

        # меняем местами столбцы в таблице Халык Банк

        data_halyk = data_halyk[
            ['Дата', 'Дата валютирования', 'Номер документа', 'Контрагент', 'БИН контрагента', 'Счет контрагента',
             'Банк контрагента', 'БИК банка контрагента', 'Дебет', 'Кредит', 'Назначение платежа']]

        # переименовываем столбцы в таблице Халык Банк

        data_halyk.columns = ['Дата', 'Дата валютирования', 'Номер', 'Контрагент', 'БИН', 'Счет контрагента',
                              'Банк контрагента', 'БИК', 'Дебет', 'Кредит', 'Назначение платежа']

    # соединяем таблицы Сбербанк и Халык Банк

    if len(data_sber) > 0:
        data = data_sber.copy()
        data = data.append(data_halyk)

    else:
        data = data_halyk


    # меняем типы данных и отсортируем таблицу

    data = data.reset_index(drop=True)

    data['БИН'] = data['БИН'].astype('int64')

    data['Дебет'] = data['Дебет'].str.replace(",", ".")
    data['Дебет'] = data['Дебет'].str.replace(" ", "")
    data['Дебет'] = data['Дебет'].astype('float64')

    data['Кредит'] = data['Кредит'].str.replace(",", ".")
    data['Кредит'] = data['Кредит'].str.replace(" ", "")
    data['Кредит'] = data['Кредит'].astype('float64')

    data = data.sort_values(by=['Дата валютирования','Номер'])

    data = data.reset_index(drop=True)

    # записывает полученные данные в файл

    data.to_excel('Итог.xlsx', sheet_name='KZ61914082203KZ017H9', index = False, engine='openpyxl')

    # открываем файл для изменения стиля ячеек

    wb = load_workbook('Итог.xlsx')
    sheet = wb['KZ61914082203KZ017H9']

    # фиксируем стиль для границ ячейки (рамку)

    bd = Side(style='thin', color="000000")

    # проводим передор ячеек и меняем стили

    col_name = 'ABCDEFGHIJK'

    for k in range(0, 11):
        for i in range(2, len(data) + 2):
            sheet[col_name[k] + str(i)].font = Font(name='Times New Roman', bold=False, size=8)
            sheet[col_name[k] + str(i)].border = Border(left=bd, top=bd, right=bd, bottom=bd)

            if k == 1:
                sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[14]

            if k == 4:
                sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[1]

            if (k == 8) or (k == 9):
                sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[4]

            if (k == 10):
                sheet[col_name[k] + str(i)].alignment = Alignment(wrap_text=True)

            if (i == 2):
                sheet[col_name[k] + str(i)].fill = PatternFill("solid", fgColor="00AAEE")

            if sheet['A' + str(i)].value == None:
                sheet[col_name[k] + str(i)].fill = PatternFill("solid", fgColor="FFFF00")

    # изменяем ширину столбцов

    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 24
    sheet.column_dimensions['H'].width = 16
    sheet.column_dimensions['I'].width = 16
    sheet.column_dimensions['J'].width = 16
    sheet.column_dimensions['K'].width = 36

    print(sheet['A2'].value)

    # записываем результат в файл

    wb.save('Итог.xlsx')

    return